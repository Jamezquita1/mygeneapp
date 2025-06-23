import openpyxl
from openpyxl.styles import Font, Alignment
from gene_functions import get_blastp
from phenotype_code import process_phenotypes
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def create_excel_file(filename):
    workbook = openpyxl.Workbook()
    default_sheet = workbook.active
    default_sheet.title = "Descriptions"

    workbook.create_sheet("Protein Homology")
    workbook.create_sheet("Phenotypes")
    workbook.create_sheet("Strain Info")

    workbook.save(filename)



def apply_header_format(cell, text):
    font = Font(bold=True)
    alignment = Alignment(horizontal='left', vertical='center')

    cell.value = text
    cell.font = font
    cell.alignment = alignment

def write_gene_data_to_excel(filename, gene_data):
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook["Descriptions"]

    # Find the next available row
    new_row = sheet.max_row + 1 if sheet.max_row > 1 else 1

    for gene_info in gene_data:
        apply_header_format(sheet.cell(row=new_row, column=1), "Gene Name:")
        sheet.cell(row=new_row, column=2, value=gene_info['gene_name'])

        apply_header_format(sheet.cell(row=new_row + 1, column=1), "WormBase Gene ID:")
        sheet.cell(row=new_row + 1, column=2, value=gene_info['gene_id'])

        apply_header_format(sheet.cell(row=new_row + 2, column=1), "Description:")
        sheet.cell(row=new_row + 2, column=2, value=gene_info['gene_description'])

        apply_header_format(sheet.cell(row=new_row + 3, column=1), "Position:")
        sheet.cell(row=new_row + 3, column=2, value=gene_info['gene_position'])

        apply_header_format(sheet.cell(row=new_row + 4, column=1), "Other Names:")
        sheet.cell(row=new_row + 4, column=2, value=gene_info['gene_othernames'])

        new_row += 6  # Leave some space before next gene

    workbook.save(filename)


def write_protein_data_to_excel(filename, protein_data):
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook["Protein Homology"]
    print("protein data = ", protein_data)

    # Add header only if sheet is empty
    if sheet.max_row == 1 and all(cell.value is None for cell in sheet[1]):
        sheet.append(["Gene Label", "Protein Label", "Protein ID", "ID", "Percentage", "E-value", "Source Range", "Target Range"])

    for protein_info in protein_data:
        protein_id = protein_info["protein_id"]
        protein_blastp = get_blastp([protein_id])

        if protein_blastp:
            for hit in protein_blastp:
                sheet.append([
                    protein_info.get('gene_label'),
                    protein_info.get('protein_label'),
                    protein_id,
                    hit.get('id'),
                    hit.get('percentage'),
                    hit.get('evalue'),
                    hit.get('source_range'),
                    hit.get('target_range'),
                ])

    workbook.save(filename)



def write_phenotype_data_to_excel(filename, gene_data):
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook["Phenotypes"]

    # Add header only if sheet is empty
    if sheet.max_row == 1 and all(cell.value is None for cell in sheet[1]):
        sheet.append(["Gene", "Phenotype", "Evidence Type", "Mutant/Alteration", "Evidence_Info"])

    for gene_info in gene_data:
        gene_id = gene_info["gene_id"]
        gene_name = gene_info["gene_name"]
        phenotypes = process_phenotypes(gene_id)

        for phenotype in phenotypes:
            sheet.append([
                gene_name,
                phenotype.get("Phenotype"),
                phenotype.get("Evidence Type"),
                phenotype.get("Mutant/Alteration"),
                str(phenotype.get("Evidence_Info")),
            ])

    workbook.save(filename)


def write_strain_data_to_excel(filename, strain_data):
    wb = load_workbook(filename)

    # Reuse existing sheet or create new one
    if "Strain Info" in wb.sheetnames:
        ws = wb["Strain Info"]
    else:
        ws = wb.create_sheet("Strain Info")

    # If sheet is empty, add headers
    if ws.max_row == 1 and all(cell.value is None for cell in ws[1]):
        headers = ['Gene Name', 'Strain Name', 'Species', 'Genotype', 'Additional Info']
        ws.append(headers)

    for entry in strain_data:
        ws.append([
            entry.get('gene_name'),
            entry.get('strain_name'),
            entry.get('species'),
            entry.get('genotype'),
            entry.get('additional_info'),
        ])

    # Optional: Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[column].width = max_length + 2

    wb.save(filename)

def write_failed_queries_to_excel(filename, failed_queries):
    workbook = openpyxl.load_workbook(filename)
    if "Unsuccessful Queries" in workbook.sheetnames:
        sheet = workbook["Unsuccessful Queries"]
    else:
        sheet = workbook.create_sheet("Unsuccessful Queries")

    if sheet.max_row == 1:  # Write headers if empty
        sheet.append(["User Input", "Error Message"])

    for input_str, error in failed_queries:
        sheet.append([input_str, error])

    workbook.save(filename)
